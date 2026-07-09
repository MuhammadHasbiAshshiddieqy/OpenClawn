"""Tests untuk Tools + Sandbox — Sprint 2."""

import dataclasses
import pytest
from unittest.mock import AsyncMock, MagicMock, patch
from tools.file_ops import FileReadTool, FileWriteTool
from tools.web import HttpRequestTool, WebFetchTool, _ssrf_guard
from tools.interaction import AskUserTool
from tools.code import CodeRunTool
from tools import TOOL_REGISTRY


def dataclasses_replace(obj, **changes):
    """Alias ringkas untuk dataclasses.replace (AppConfig frozen → buat salinan)."""
    return dataclasses.replace(obj, **changes)


# ── TOOL_REGISTRY ─────────────────────────────────────────────────────────────


def test_registry_has_all_27_tools():
    """Semua 27 tool harus terdaftar di TOOL_REGISTRY."""
    expected = {
        "file_read",
        "read_many",
        "file_write",
        "file_edit",
        "file_append",
        "apply_patch",
        "list_dir",
        "set_workdir",
        "glob",
        "grep",
        "pdf_read",
        "doc_write",
        "pdf_write",
        "git_status",
        "git_diff",
        "git_log",
        "shell_run",
        "code_run",
        "web_fetch",
        "web_search",
        "http_request",
        "db_query",
        "memory_search",
        "json_query",
        "ask_user",
        "todo_write",
        "report_blocker",
    }
    assert set(TOOL_REGISTRY.keys()) == expected


def test_code_run_requires_approval():
    """code_run HARUS requires_approval=True — keamanan wajib."""
    assert TOOL_REGISTRY["code_run"].requires_approval is True


def test_file_write_requires_approval():
    """file_write HARUS requires_approval=True — tool destruktif (modifikasi filesystem)."""
    assert TOOL_REGISTRY["file_write"].requires_approval is True


def test_non_destructive_tools_no_approval():
    """file_read, web_fetch, ask_user tidak butuh approval."""
    for name in ("file_read", "web_fetch", "ask_user"):
        assert TOOL_REGISTRY[name].requires_approval is False, f"{name} seharusnya False"


def test_all_destructive_tools_require_approval():
    """Semua tool yang memodifikasi state (code_run, file_write) harus requires_approval=True."""
    destructive = [n for n, t in TOOL_REGISTRY.items() if t.requires_approval]
    assert "code_run" in destructive
    assert "file_write" in destructive


def test_all_tools_have_schema():
    """Semua tool harus bisa produce schema dict yang valid."""
    for name, tool in TOOL_REGISTRY.items():
        schema = tool.schema()
        assert "name" in schema, f"{name}: schema harus punya 'name'"
        assert "input_schema" in schema, f"{name}: schema harus punya 'input_schema'"
        assert schema["name"] == name


# ── FileReadTool ──────────────────────────────────────────────────────────────


@pytest.mark.asyncio
def _set_workspace(monkeypatch, path):
    """Arahkan workspace_root tool file ke `path` (CONFIG frozen → ganti referensi)."""
    import dataclasses

    from infra.config import CONFIG

    patched = dataclasses.replace(CONFIG, workspace_root=str(path))
    monkeypatch.setattr("tools.file_ops.CONFIG", patched)


async def test_file_read_success(tmp_path, monkeypatch):
    """file_read harus mengembalikan isi file (dalam workspace)."""
    _set_workspace(monkeypatch, tmp_path)
    f = tmp_path / "test.txt"
    f.write_text("hello world")

    tool = FileReadTool()
    result = await tool.execute({"path": "test.txt"}, vault=None)
    assert result["content"] == "hello world"


@pytest.mark.asyncio
async def test_file_read_not_found():
    """file_read harus mengembalikan error jika file tidak ada."""
    tool = FileReadTool()
    result = await tool.execute({"path": "/tidak/ada/file.txt"}, vault=None)
    assert "error" in result


@pytest.mark.asyncio
async def test_file_read_no_path():
    """file_read tanpa path harus return error, tidak crash."""
    tool = FileReadTool()
    result = await tool.execute({}, vault=None)
    assert "error" in result


@pytest.mark.asyncio
async def test_file_read_truncates_large_file(tmp_path, monkeypatch):
    """file_read harus truncate konten > 10000 karakter."""
    _set_workspace(monkeypatch, tmp_path)
    f = tmp_path / "big.txt"
    f.write_text("x" * 20000)

    tool = FileReadTool()
    result = await tool.execute({"path": "big.txt"}, vault=None)
    assert len(result["content"]) <= 10000


# ── FileWriteTool ─────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_file_write_success(tmp_path, monkeypatch):
    """file_write harus menulis konten dan mengembalikan ok=True."""
    _set_workspace(monkeypatch, tmp_path)
    tool = FileWriteTool()
    result = await tool.execute({"path": "output.txt", "content": "isi file"}, vault=None)
    assert result["ok"] is True
    assert (tmp_path / "output.txt").read_text() == "isi file"


@pytest.mark.asyncio
async def test_file_write_no_path():
    """file_write tanpa path harus return error."""
    tool = FileWriteTool()
    result = await tool.execute({"content": "isi"}, vault=None)
    assert "error" in result


# ── WebFetchTool ──────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_web_fetch_success():
    """web_fetch berhasil → mengembalikan status dan konten."""
    tool = WebFetchTool()

    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.text = "page content"

    mock_client = AsyncMock()
    mock_client.__aenter__ = AsyncMock(return_value=mock_client)
    mock_client.__aexit__ = AsyncMock(return_value=False)
    mock_client.get = AsyncMock(return_value=mock_resp)

    # _ssrf_guard di-bypass (return None = aman) agar test tak melakukan DNS nyata.
    with (
        patch("tools.web.httpx.AsyncClient", return_value=mock_client),
        patch("tools.web._ssrf_guard", return_value=None),
    ):
        result = await tool.execute({"url": "http://example.com"}, vault=None)

    assert result["status"] == 200
    assert "page content" in result["content"]


@pytest.mark.asyncio
async def test_web_fetch_no_url():
    """web_fetch tanpa url harus return error."""
    tool = WebFetchTool()
    result = await tool.execute({}, vault=None)
    assert "error" in result


@pytest.mark.asyncio
async def test_web_fetch_http_error():
    """web_fetch dengan HTTP error harus return error, tidak raise."""
    import httpx

    tool = WebFetchTool()

    mock_client = AsyncMock()
    mock_client.__aenter__ = AsyncMock(return_value=mock_client)
    mock_client.__aexit__ = AsyncMock(return_value=False)
    mock_client.get = AsyncMock(side_effect=httpx.HTTPError("connection error"))

    with (
        patch("tools.web.httpx.AsyncClient", return_value=mock_client),
        patch("tools.web._ssrf_guard", return_value=None),
    ):
        result = await tool.execute({"url": "https://bad-url.example"}, vault=None)

    assert "error" in result


# ── SSRF guard (web_fetch + http_request) ─────────────────────────────────────


def test_ssrf_guard_blocks_loopback():
    """localhost / 127.0.0.1 ditolak (IP literal → tanpa DNS nyata)."""
    assert _ssrf_guard("http://127.0.0.1:11434/api") is not None
    assert _ssrf_guard("http://[::1]/x") is not None


def test_ssrf_guard_blocks_cloud_metadata():
    """Endpoint metadata cloud (link-local 169.254.169.254) ditolak."""
    blocked = _ssrf_guard("http://169.254.169.254/latest/meta-data/")
    assert blocked is not None and "SSRF" in blocked


def test_ssrf_guard_blocks_private_rfc1918():
    """Alamat privat RFC1918 (mis. 10.x, 192.168.x) ditolak."""
    assert _ssrf_guard("http://10.0.0.5/") is not None
    assert _ssrf_guard("http://192.168.1.1/admin") is not None


def test_ssrf_guard_allows_public_ip():
    """IP publik (literal) lolos guard."""
    assert _ssrf_guard("http://8.8.8.8/") is None


def test_ssrf_guard_blocks_dns_rebinding(monkeypatch):
    """Nama domain yang resolve ke IP internal tetap diblokir (bukan hanya literal IP)."""

    def fake_getaddrinfo(host, port, **kwargs):
        # Simulasikan domain jahat yang menunjuk ke loopback.
        return [(2, 1, 6, "", ("127.0.0.1", port or 80))]

    monkeypatch.setattr("tools.web.socket.getaddrinfo", fake_getaddrinfo)
    assert _ssrf_guard("http://evil.example.com/") is not None


@pytest.mark.asyncio
async def test_web_fetch_rejects_internal_host():
    """web_fetch ke host internal ditolak SEBELUM request keluar (tanpa approval)."""
    tool = WebFetchTool()
    result = await tool.execute({"url": "http://localhost:11434/api/tags"}, vault=None)
    assert "error" in result and "SSRF" in result["error"]


@pytest.mark.asyncio
async def test_web_fetch_rejects_non_http_scheme():
    """Scheme selain http/https (mis. file://) ditolak."""
    tool = WebFetchTool()
    result = await tool.execute({"url": "file:///etc/passwd"}, vault=None)
    assert "error" in result


@pytest.mark.asyncio
async def test_http_request_rejects_internal_host():
    """http_request juga diblokir SSRF walau butuh approval (approval bukan satu-satunya penghalang)."""
    tool = HttpRequestTool()
    result = await tool.execute(
        {"url": "http://169.254.169.254/latest/meta-data/", "method": "GET"},
        vault=None,
    )
    assert "error" in result and "SSRF" in result["error"]


# ── AskUserTool ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_ask_user_returns_stub():
    """ask_user stub harus mengembalikan jawaban tanpa crash."""
    tool = AskUserTool()
    result = await tool.execute({"question": "apa preferensimu?"}, vault=None)
    assert "answer" in result


@pytest.mark.asyncio
async def test_ask_user_no_question():
    """ask_user tanpa question harus tetap return dict tanpa crash."""
    tool = AskUserTool()
    result = await tool.execute({}, vault=None)
    assert isinstance(result, dict)


# ── CodeRunTool + DockerSandbox ───────────────────────────────────────────────


@pytest.mark.asyncio
async def test_code_run_no_code():
    """code_run tanpa kode harus return error, tidak jalankan Docker."""
    tool = CodeRunTool()
    result = await tool.execute({}, vault=None)
    assert "error" in result


@pytest.mark.asyncio
async def test_code_run_delegates_to_sandbox():
    """code_run harus mendelegasikan eksekusi ke DockerSandbox, bukan langsung exec."""
    tool = CodeRunTool()
    tool.sandbox.run_python = AsyncMock(return_value={"stdout": "42\n", "exit_code": 0})

    result = await tool.execute({"code": "print(42)"}, vault=None)
    tool.sandbox.run_python.assert_called_once_with("print(42)")
    assert result["stdout"] == "42\n"


@pytest.mark.asyncio
async def test_sandbox_timeout_handled():
    """Sandbox timeout harus return error dict, tidak raise ke caller."""
    from tools.sandbox import DockerSandbox
    import asyncio

    sandbox = DockerSandbox()

    async def _fake_exec(*args, **kwargs):
        raise asyncio.TimeoutError()

    with patch("tools.sandbox.asyncio.create_subprocess_exec", side_effect=asyncio.TimeoutError):
        result = await sandbox.run_python("import time; time.sleep(999)")

    assert "error" in result
    assert result["exit_code"] == -1


def _flag_pair_present(argv: list[str], flag: str, value: str | None) -> bool:
    """True bila `flag` ada di argv dan (jika value diberikan) diikuti value tepat."""
    if flag not in argv:
        return False
    if value is None:
        return True
    i = argv.index(flag)
    return i + 1 < len(argv) and argv[i + 1] == value


async def _capture_docker_argv(coro_factory) -> list[str]:
    """Jalankan satu metode sandbox & tangkap argv NYATA yang dikirim ke Docker.

    Bukan rekonstruksi manual — kita patch create_subprocess_exec dan rekam
    argumen sesungguhnya, sehingga test gagal bila kode menghapus flag keamanan.
    """
    captured: dict = {}

    async def _fake_exec(*args, **kwargs):
        captured["argv"] = list(args)
        proc = MagicMock()
        proc.communicate = AsyncMock(return_value=(b"", b""))
        proc.returncode = 0
        return proc

    with patch("tools.sandbox.asyncio.create_subprocess_exec", side_effect=_fake_exec):
        await coro_factory()
    return captured["argv"]


@pytest.mark.asyncio
async def test_run_python_argv_enforces_security_flags():
    """run_python HARUS mengirim --network none, --read-only, non-root, no-new-privileges.

    Memeriksa argv NYATA (anti test palsu): hapus salah satu flag di sandbox.py
    → test ini gagal. Mount /work WAJIB read-only (:ro).
    """
    from tools.sandbox import DockerSandbox

    sandbox = DockerSandbox()
    argv = await _capture_docker_argv(lambda: sandbox.run_python("print(1)"))

    assert _flag_pair_present(argv, "--network", "none"), "Harus --network none"
    assert _flag_pair_present(argv, "--read-only", None), "Harus --read-only"
    assert _flag_pair_present(argv, "--user", "nobody"), "Harus user non-root"
    assert _flag_pair_present(argv, "--security-opt", "no-new-privileges"), (
        "Harus no-new-privileges"
    )
    # Tidak boleh ada mount writable ke /work — semua mount -v harus berakhiran :ro
    mounts = [argv[i + 1] for i, a in enumerate(argv) if a == "-v"]
    assert mounts and all(m.endswith(":ro") for m in mounts), (
        f"Semua mount -v harus read-only, dapat: {mounts}"
    )


@pytest.mark.asyncio
async def test_run_shell_argv_enforces_security_flags():
    """run_shell HARUS menegakkan flag keamanan yang sama + mount workspace read-only."""
    from tools.sandbox import DockerSandbox

    sandbox = DockerSandbox()
    argv = await _capture_docker_argv(lambda: sandbox.run_shell("ls", workspace_root="."))

    assert _flag_pair_present(argv, "--network", "none"), "Harus --network none"
    assert _flag_pair_present(argv, "--read-only", None), "Harus --read-only"
    assert _flag_pair_present(argv, "--user", "nobody"), "Harus user non-root"
    assert _flag_pair_present(argv, "--security-opt", "no-new-privileges"), (
        "Harus no-new-privileges"
    )
    mounts = [argv[i + 1] for i, a in enumerate(argv) if a == "-v"]
    assert mounts and all(m.endswith(":ro") for m in mounts), "Workspace mount harus :ro"


@pytest.mark.asyncio
async def test_run_python_fails_safe_when_docker_absent():
    """Docker tak terpasang → SandboxUnavailable, BUKAN fallback eksekusi di host.

    Prinsip keamanan #1: tidak pernah ada eksekusi kode di host.
    """
    from tools.sandbox import DockerSandbox, SandboxUnavailable

    sandbox = DockerSandbox()
    with patch(
        "tools.sandbox.asyncio.create_subprocess_exec",
        side_effect=FileNotFoundError("docker not found"),
    ):
        with pytest.raises(SandboxUnavailable):
            await sandbox.run_python("print(1)")


@pytest.mark.asyncio
async def test_run_shell_fails_safe_when_docker_absent():
    """Sama untuk run_shell — Docker absen harus fail-safe, tidak jatuh ke host."""
    from tools.sandbox import DockerSandbox, SandboxUnavailable

    sandbox = DockerSandbox()
    with patch(
        "tools.sandbox.asyncio.create_subprocess_exec",
        side_effect=FileNotFoundError("docker not found"),
    ):
        with pytest.raises(SandboxUnavailable):
            await sandbox.run_shell("ls", workspace_root=".")


def test_base_docker_args_contains_every_required_flag():
    """_base_docker_args (sumber tunggal argv) harus memuat SEMUA _REQUIRED_FLAGS.

    Guard tambahan di level builder: kalau flag dihapus dari sumbernya, test gagal.
    """
    from tools.sandbox import DockerSandbox, _REQUIRED_FLAGS

    args = DockerSandbox()._base_docker_args("/x:/work:ro", "16m")
    for pair in _REQUIRED_FLAGS:
        flag = pair[0]
        value = pair[1] if len(pair) > 1 else None
        assert _flag_pair_present(args, flag, value), f"flag wajib hilang: {pair}"


# ── Approval gate integration ────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_approval_called_for_destructive_tool():
    """Tool requires_approval=True memicu HITL: request() dicatat & menunggu keputusan.

    Approval interaktif (bukan auto-approve). Di sini kita simulasikan
    user menekan 'approve' lewat resolve(). Coverage HITL lengkap di test_security.py.
    """
    import asyncio
    from security.approval import ApprovalGate
    from infra.config import AppConfig
    from infra.database import DatabaseManager

    cfg = AppConfig(db_path=":memory:", approval_timeout_sec=1)
    db = DatabaseManager(cfg)
    with open("migrations/001_initial.sql") as f:
        conn = await db.conn()
        await conn.executescript(f.read())
        await conn.commit()

    gate = ApprovalGate(db=db, config=cfg)

    async def _user_approves():
        await asyncio.sleep(0.05)
        pending = gate.pending_list("test-s1")
        gate.resolve(pending[0]["approval_id"], True)

    asyncio.create_task(_user_approves())
    approved = await gate.request(
        session_id="test-s1", tool_name="code_run", tool_input={"code": "print(42)"}
    )
    assert approved is True

    # Verifikasi tersimpan di approval_log dengan keputusan final
    row = await db.fetchone(
        "SELECT tool_name, decision FROM approval_log WHERE session_id='test-s1'"
    )
    assert row is not None
    assert row["tool_name"] == "code_run"
    assert row["decision"] == "approved"

    await db.close()


@pytest.mark.asyncio
async def test_approval_log_contains_tool_input():
    """Approval log harus menyimpan tool_input untuk audit trail.

    Regresi performa: request() TIDAK di-resolve di sini (tak ada `.resolve()`
    dipanggil), jadi tadinya menunggu penuh approval_timeout_sec DEFAULT (120s)
    sebelum fail-safe timeout mengembalikan kontrol — bikin test ini (dan suite
    penuh) makan waktu 2 menit ekstra, terlihat seperti hang padahal cuma lambat.
    approval_timeout_sec di-set sangat kecil karena test ini hanya perlu baris
    ter-insert SEGERA saat request() dipanggil, bukan menunggu siklus timeout.
    """
    from security.approval import ApprovalGate
    from infra.config import AppConfig
    from infra.database import DatabaseManager
    import json

    cfg = AppConfig(db_path=":memory:", approval_timeout_sec=1)
    db = DatabaseManager(cfg)
    with open("migrations/001_initial.sql") as f:
        conn = await db.conn()
        await conn.executescript(f.read())
        await conn.commit()

    gate = ApprovalGate(db=db, config=cfg)
    await gate.request(
        session_id="s2",
        tool_name="file_write",
        tool_input={"path": "/tmp/test.py", "content": "print(1)"},
    )

    row = await db.fetchone("SELECT tool_input FROM approval_log WHERE session_id='s2'")
    parsed = json.loads(row["tool_input"])
    assert parsed["path"] == "/tmp/test.py"

    await db.close()


# ── Jaring pengaman eksekusi tool + validasi + telemetri (_execute_tool) ──────


async def _agent_with_db():
    """AgentLoop role 'dev' (mengizinkan banyak tool) + DB :memory: siap pakai."""
    from core.agent_loop import AgentLoop, AgentConfig
    from infra.config import AppConfig
    from infra.database import DatabaseManager

    cfg = AppConfig(db_path=":memory:", tool_timeout_sec=1)
    db = DatabaseManager(cfg)
    with open("migrations/001_initial.sql") as f:
        conn = await db.conn()
        await conn.executescript(f.read())
        await conn.commit()
    agent = AgentLoop(AgentConfig(role="dev", session_id="s-tool"), db=db, config=cfg)
    return agent, db


@pytest.mark.asyncio
async def test_tool_exception_returns_error_not_crash():
    """Tool yang melempar exception → error dict anggun (§1.3), turn tidak mati."""
    agent, db = await _agent_with_db()
    tool = TOOL_REGISTRY["grep"]
    with patch.object(tool, "execute", side_effect=RuntimeError("boom")):
        result = await agent._execute_tool("grep", {"pattern": "x"})
    assert "error" in result
    assert "boom" in result["error"]
    await db.close()


@pytest.mark.asyncio
async def test_tool_timeout_returns_error():
    """Tool yang menggantung meldebihi tool_timeout_sec → error timeout, bukan freeze."""
    import asyncio

    agent, db = await _agent_with_db()

    async def _hang(*a, **k):
        await asyncio.sleep(10)

    tool = TOOL_REGISTRY["grep"]
    with patch.object(tool, "execute", side_effect=_hang):
        result = await agent._execute_tool("grep", {"pattern": "x"})
    assert "error" in result and "waktu" in result["error"]
    await db.close()


@pytest.mark.asyncio
async def test_tool_output_truncated_uniformly():
    """Output teks panjang dipotong ke tool_max_output, apa pun tool-nya."""
    agent, db = await _agent_with_db()
    agent.config = dataclasses_replace(agent.config, tool_max_output=50)
    huge = {"content": "a" * 5000}
    tool = TOOL_REGISTRY["grep"]
    with patch.object(tool, "execute", new=AsyncMock(return_value=huge)):
        result = await agent._execute_tool("grep", {"pattern": "x"})
    assert len(result["content"]) < 200
    assert "dipotong" in result["content"]
    await db.close()


@pytest.mark.asyncio
async def test_tool_missing_required_field_rejected_before_execute():
    """Field required hilang → error jelas, tool TIDAK dieksekusi (validasi schema)."""
    agent, db = await _agent_with_db()
    tool = TOOL_REGISTRY["grep"]
    called = {"n": 0}

    async def _spy(*a, **k):
        called["n"] += 1
        return {}

    with patch.object(tool, "execute", side_effect=_spy):
        result = await agent._execute_tool("grep", {})  # 'pattern' wajib, hilang
    assert "error" in result and "pattern" in result["error"]
    assert called["n"] == 0  # tidak dieksekusi
    await db.close()


@pytest.mark.asyncio
async def test_tool_invocation_recorded_in_telemetry():
    """Eksekusi tool tercatat di tool_invocations dengan outcome & latency."""
    agent, db = await _agent_with_db()
    tool = TOOL_REGISTRY["grep"]
    with patch.object(tool, "execute", new=AsyncMock(return_value={"matches": []})):
        await agent._execute_tool("grep", {"pattern": "x"})
    row = await db.fetchone(
        "SELECT tool_name, outcome, latency_ms FROM tool_invocations WHERE tool_name='grep'"
    )
    assert row is not None
    assert row["outcome"] == "ok"
    assert row["latency_ms"] is not None
    await db.close()


@pytest.mark.asyncio
async def test_tool_failure_recorded_as_error_outcome():
    """Tool gagal → telemetri mencatat outcome='error'."""
    agent, db = await _agent_with_db()
    tool = TOOL_REGISTRY["grep"]
    with patch.object(tool, "execute", side_effect=ValueError("x")):
        await agent._execute_tool("grep", {"pattern": "x"})
    row = await db.fetchone("SELECT outcome FROM tool_invocations WHERE tool_name='grep'")
    assert row["outcome"] == "error"
    await db.close()


@pytest.mark.asyncio
async def test_tool_audit_summary_aggregates():
    """ToolAudit.summary() mengagregasi total/errors/fail_rate per tool."""
    from core.tool_audit import ToolAudit

    agent, db = await _agent_with_db()
    audit = ToolAudit(db)
    await audit.record("s", "dev", "grep", "ok", 10)
    await audit.record("s", "dev", "grep", "error", 20)
    await audit.record("s", "dev", "glob", "ok", 5)
    summary = await audit.summary()
    by_tool = {r["tool_name"]: r for r in summary}
    assert by_tool["grep"]["total"] == 2
    assert by_tool["grep"]["errors"] == 1
    assert by_tool["grep"]["fail_rate"] == 50.0
    await db.close()


@pytest.mark.asyncio
async def test_tool_audit_record_defaults_actor_is_agent_true():
    """Audit log format actor_is_agent (TODO.md § Prioritas 2): semua baris
    tool_invocations adalah tindakan agent, bukan manusia langsung."""
    from core.tool_audit import ToolAudit

    agent, db = await _agent_with_db()
    audit = ToolAudit(db)
    await audit.record("s", "dev", "grep", "ok", 10)

    row = await db.fetchone("SELECT actor_is_agent FROM tool_invocations WHERE tool_name='grep'")
    assert row["actor_is_agent"] == 1
    await db.close()


@pytest.mark.asyncio
async def test_tool_audit_record_stores_user_id():
    """user_id opsional (default 'default') query-able terpisah dari session_id
    untuk integrasi SIEM eksternal."""
    from core.tool_audit import ToolAudit

    agent, db = await _agent_with_db()
    audit = ToolAudit(db)
    await audit.record("s", "dev", "grep", "ok", 10, user_id="bob")

    row = await db.fetchone("SELECT user_id FROM tool_invocations WHERE tool_name='grep'")
    assert row["user_id"] == "bob"
    await db.close()


# ── read_many (batch file read) ──────────────────────────────────────────────


def _patch_workspace(monkeypatch, tmp_path, *mods):
    """CONFIG frozen → ganti referensi CONFIG modul dengan salinan ber-workspace tmp_path."""
    import dataclasses
    from infra.config import CONFIG

    patched = dataclasses.replace(CONFIG, workspace_root=str(tmp_path))
    for mod in mods:
        monkeypatch.setattr(f"{mod}.CONFIG", patched)


@pytest.mark.asyncio
async def test_read_many_reads_multiple_files(tmp_path, monkeypatch):
    """read_many membaca beberapa file dalam satu panggilan, semua workspace-safe."""
    from tools.file_ops import ReadManyTool

    (tmp_path / "a.txt").write_text("isi-a")
    (tmp_path / "b.txt").write_text("isi-b")
    _patch_workspace(monkeypatch, tmp_path, "tools.file_ops")

    result = await ReadManyTool().execute({"paths": ["a.txt", "b.txt"]}, vault=None)
    assert result["count"] == 2
    contents = {f["path"]: f.get("content") for f in result["files"]}
    assert contents["a.txt"] == "isi-a"
    assert contents["b.txt"] == "isi-b"


@pytest.mark.asyncio
async def test_read_many_per_file_error_does_not_fail_others(tmp_path, monkeypatch):
    """Satu file hilang → error per-file, file lain tetap terbaca (tidak crash)."""
    from tools.file_ops import ReadManyTool

    (tmp_path / "ok.txt").write_text("ada")
    _patch_workspace(monkeypatch, tmp_path, "tools.file_ops")

    result = await ReadManyTool().execute({"paths": ["ok.txt", "hilang.txt"]}, vault=None)
    files = {f["path"]: f for f in result["files"]}
    assert files["ok.txt"]["content"] == "ada"
    assert "error" in files["hilang.txt"]


@pytest.mark.asyncio
async def test_read_many_requires_list():
    """paths bukan list → error, tidak crash."""
    from tools.file_ops import ReadManyTool

    result = await ReadManyTool().execute({"paths": "bukan-list"}, vault=None)
    assert "error" in result


@pytest.mark.asyncio
async def test_read_many_caps_batch_size(tmp_path, monkeypatch):
    """Lebih dari MAX_FILES_PER_BATCH → dipotong, skipped dilaporkan."""
    from tools.file_ops import ReadManyTool, MAX_FILES_PER_BATCH

    _patch_workspace(monkeypatch, tmp_path, "tools.file_ops")
    paths = [f"f{i}.txt" for i in range(MAX_FILES_PER_BATCH + 5)]
    result = await ReadManyTool().execute({"paths": paths}, vault=None)
    assert result["count"] == MAX_FILES_PER_BATCH
    assert result["skipped"] == 5


# ── doc_write (docx/pptx/xlsx/md) ────────────────────────────────────────────


@pytest.mark.asyncio
async def test_doc_write_requires_approval():
    """doc_write menulis file → harus requires_approval=True."""
    assert TOOL_REGISTRY["doc_write"].requires_approval is True


@pytest.mark.asyncio
async def test_doc_write_rejects_unknown_format(tmp_path, monkeypatch):
    """Format tak dikenal → error, tidak menulis apa pun."""
    from tools.document import DocWriteTool

    _patch_workspace(monkeypatch, tmp_path, "tools.document")
    result = await DocWriteTool().execute(
        {"path": "x.foo", "format": "foo", "content": "hi"}, vault=None
    )
    assert "error" in result


@pytest.mark.asyncio
async def test_doc_write_markdown_string(tmp_path, monkeypatch):
    """md dengan content string → file teks tertulis."""
    from tools.document import DocWriteTool

    _patch_workspace(monkeypatch, tmp_path, "tools.document")
    result = await DocWriteTool().execute(
        {"path": "out.md", "format": "md", "content": "# Judul\n\nisi"}, vault=None
    )
    assert result.get("ok") is True
    assert (tmp_path / "out.md").read_text().startswith("# Judul")


@pytest.mark.asyncio
async def test_doc_write_docx_structured(tmp_path, monkeypatch):
    """docx dari {title, sections} → file .docx valid yang bisa dibuka kembali."""
    from tools.document import DocWriteTool
    from docx import Document

    _patch_workspace(monkeypatch, tmp_path, "tools.document")
    content = {
        "title": "Laporan",
        "sections": [{"heading": "Ringkasan", "body": "teks", "bullets": ["a", "b"]}],
    }
    result = await DocWriteTool().execute(
        {"path": "r.docx", "format": "docx", "content": content}, vault=None
    )
    assert result.get("ok") is True
    d = Document(str(tmp_path / "r.docx"))
    texts = [p.text for p in d.paragraphs]
    assert "Laporan" in texts and "teks" in texts


@pytest.mark.asyncio
async def test_doc_write_xlsx_rows(tmp_path, monkeypatch):
    """xlsx dari {headers, rows} → spreadsheet dengan baris benar."""
    from tools.document import DocWriteTool
    from openpyxl import load_workbook

    _patch_workspace(monkeypatch, tmp_path, "tools.document")
    content = {"headers": ["Nama", "Skor"], "rows": [["A", 1], ["B", 2]]}
    result = await DocWriteTool().execute(
        {"path": "data.xlsx", "format": "xlsx", "content": content}, vault=None
    )
    assert result.get("ok") is True
    wb = load_workbook(str(tmp_path / "data.xlsx"))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == ("Nama", "Skor")
    assert rows[1] == ("A", 1)


@pytest.mark.asyncio
async def test_doc_write_pptx_slides(tmp_path, monkeypatch):
    """pptx dari {title, slides} → presentasi dengan slide judul + konten."""
    from tools.document import DocWriteTool
    from pptx import Presentation

    _patch_workspace(monkeypatch, tmp_path, "tools.document")
    content = {"title": "Deck", "slides": [{"title": "Slide 1", "bullets": ["poin"]}]}
    result = await DocWriteTool().execute(
        {"path": "deck.pptx", "format": "pptx", "content": content}, vault=None
    )
    assert result.get("ok") is True
    prs = Presentation(str(tmp_path / "deck.pptx"))
    assert len(prs.slides) >= 2  # judul + 1 konten


@pytest.mark.asyncio
async def test_doc_write_rejects_path_outside_workspace(tmp_path, monkeypatch):
    """Path di luar workspace ditolak (keamanan #1)."""
    from tools.document import DocWriteTool

    _patch_workspace(monkeypatch, tmp_path, "tools.document")
    result = await DocWriteTool().execute(
        {"path": "../escape.md", "format": "md", "content": "x"}, vault=None
    )
    assert "error" in result


# ── Akses doc_write/pdf_write per role (§ user request: QA harus bisa menulis
# test-case matrix Excel / laporan PDF, bukan hanya file teks via file_write) ──


@pytest.mark.parametrize("role", ["pm", "dev", "qa"])
def test_pm_dev_qa_have_doc_and_pdf_write_access(role):
    """pm/dev/qa punya doc_write+pdf_write di [tools].allowed soul.toml.

    Sebelumnya qa hanya punya file_write — permintaan format office/PDF (test-case
    matrix sebagai Excel, laporan bug sebagai PDF) tak punya tool yang tepat.
    """
    import tomllib

    with open(f"roles/{role}/soul.toml", "rb") as f:
        soul = tomllib.load(f)
    allowed = soul["tools"]["allowed"]
    assert "doc_write" in allowed, f"{role} soul.toml belum punya doc_write"
    assert "pdf_write" in allowed, f"{role} soul.toml belum punya pdf_write"


def test_security_soul_unchanged_read_only():
    """Kontrol negatif: role security TETAP tidak punya akses tulis (read-only §17)."""
    import tomllib

    with open("roles/security/soul.toml", "rb") as f:
        soul = tomllib.load(f)
    allowed = soul["tools"]["allowed"]
    assert "doc_write" not in allowed
    assert "pdf_write" not in allowed
    assert "file_write" not in allowed
